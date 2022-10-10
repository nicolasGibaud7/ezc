from typing import Any, List

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, Side

from ezc.constants import (
    EXCEL_COLUMNS,
    FIRST_INGREDIENT_ROW,
    RECIPE_NAME_EXCEL_CASE,
    TITLE_CELL,
)
from ezc.recipes import Ingredient, RecipeElement
from ezc.shopping import ShoppingList


class ExcelFactory:
    """
    Represent a Excel files with one or several sheets and
    provide a set of tools to interact with Excel Sheets.
    """

    def __init__(self, name: str) -> None:
        self.name = name
        try:
            self.workbook = load_workbook(
                name
            )  # TODO Improve the searching method -> Recursive search into directories
        except FileNotFoundError:
            self.workbook = Workbook()
            self.workbook.save(filename=name)
        self.current_sheet = self.workbook.active

    def add_element(self, element_information: List[str]):
        """Add an element information to excel spreadsheet at the first row

        Args:
            element_information (List[str]): List of element information
        """
        for index, information in enumerate(element_information):
            cell = self.current_sheet[f"{EXCEL_COLUMNS[index]}1"]
            cell.value = information
        self.workbook.save(self.name)

    def add_shopping_list(self, shopping_list: ShoppingList):
        for index, element in enumerate(shopping_list.elements):
            self.add_ingredient(
                [
                    element.ingredient.name,
                    element.ingredient.shelf,
                    str(element.quantity),
                    element.ingredient.unite,
                    str(element.price),
                ],
                4 + index,
            )

    def add_ingredient(
        self, ingredient_information: List[str], row_index: int
    ):
        """Add ingredient information at a given row to excel spreadsheet

        Args:
            ingredient_information (List[str]): List of all ingredient information
            row_index (int): Row index where all ingredient information will be added
        """
        for column_index, information in enumerate(ingredient_information):
            cell = self.current_sheet[
                f"{EXCEL_COLUMNS[column_index]}{row_index}"
            ]
            cell.value = information
        self.workbook.save(self.name)

    def add_title(self):
        """Formatize the cell at the TITLE_CELL position to a title style:
        - Thin borders
        - Center alignment
        - Bold font
        And add the name of the file as the title value.
        """
        title_cell = self.current_sheet[TITLE_CELL]
        self._add_title_style(title_cell)
        title_cell.value = self.name.split(".")[0].capitalize()

    def create_header_row(self, categories: List[str]):
        """Formatize a row to a specific header style and add the several
        categories as header titles

        Args:
            categories (List[str]): Header titles
        """
        for index, header_column in enumerate(
            EXCEL_COLUMNS[0 : len(categories)]
        ):
            header_cell = self.current_sheet[f"{header_column}3"]
            self._add_header_style(header_cell)
            header_cell.value = categories[index]

    def get_recipe_name(self) -> str:
        """Get the recipe name on recipe description spreadsheet

        Returns:
            str: Recipe name
        """
        return self.current_sheet[RECIPE_NAME_EXCEL_CASE].value

    def iterate_ingredient(self):
        for ingredient_attr in self._iterate_ingredients():
            yield Ingredient(
                ingredient_attr[0].value,
                ingredient_attr[1].value,
                ingredient_attr[2].value,
                ingredient_attr[3].value,
            )

    def iterate_recipe_element(self):
        for recipe_element_attr in self._iterate_ingredients():
            yield RecipeElement(
                recipe_element_attr[0].value, recipe_element_attr[1].value
            )

    def _iterate_ingredients(self):
        """Iterate on ingredients row on recipe and ingredients file

        Returns:
            Generator: Generator on the current ingredient row
        """
        return self.current_sheet.iter_rows(min_row=FIRST_INGREDIENT_ROW)

    def _add_title_style(self, cell: Cell):
        cell.font = Font(bold=True)
        thin_side = Side(border_style="thin")
        cell.border = Border(
            bottom=thin_side, top=thin_side, left=thin_side, right=thin_side
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _add_header_style(self, cell: Cell):
        cell.style = "Headline 2"
